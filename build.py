"""
CTR 전결검색 HTML 빌더 (다중 엑셀 지원)
data/ 폴더의 모든 .xlsx 파일 → 탭 형태로 한 페이지에 통합 → AES 암호화.
"""
from __future__ import annotations
 
import base64
import glob
import json
import os
import re
import sys
from pathlib import Path
 
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.ciphers import Cipher, algorithms, modes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from openpyxl import load_workbook
 
ROOT = Path(__file__).parent
DATA_DIR = ROOT / "data"
WRAPPER_TPL = ROOT / "template" / "wrapper.html"
INNER_TPL = ROOT / "template" / "inner.html"
OUTPUT = ROOT / "CTR_전결검색.html"
ITERATIONS = 600_000
 
 
def find_excels():
    return [Path(p) for p in sorted(glob.glob(str(DATA_DIR / "*.xlsx")))]
 
 
def read_excel(path):
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    body = []
    for row in rows:
        cells = [("" if v is None else str(v)) for v in row[: len(header)]]
        if any(c.strip() for c in cells):
            body.append(cells)
    wb.close()
    return header, body
 
 
def tab_label(name):
    stem = Path(name).stem
    cleaned = re.sub(r"[_\-\s]*(rev|ver|v)\.?\d.*$", "", stem, flags=re.IGNORECASE)
    cleaned = re.sub(r"[_\-\s]*\d{4}[\-_\.]?\d{2}[\-_\.]?\d{2}.*$", "", cleaned)
    return cleaned or stem
 
 
def encrypt(plaintext, password):
    salt = os.urandom(16)
    iv = os.urandom(16)
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=ITERATIONS)
    key = kdf.derive(password.encode("utf-8"))
    data = plaintext.encode("utf-8")
    pad_len = 16 - (len(data) % 16)
    data += bytes([pad_len]) * pad_len
    cipher = Cipher(algorithms.AES(key), modes.CBC(iv))
    enc = cipher.encryptor()
    ct = enc.update(data) + enc.finalize()
    return {
        "salt": base64.b64encode(salt).decode("ascii"),
        "iv": base64.b64encode(iv).decode("ascii"),
        "ciphertext": base64.b64encode(ct).decode("ascii"),
        "iterations": ITERATIONS,
    }
 
 
def main():
    password = os.environ.get("PASSWORD")
    if not password:
        sys.exit("ERROR: PASSWORD 환경변수가 없습니다.")
    if not WRAPPER_TPL.exists():
        sys.exit(f"ERROR: 래퍼 템플릿 없음: {WRAPPER_TPL}")
    if not INNER_TPL.exists():
        sys.exit(f"ERROR: 내부 템플릿 없음: {INNER_TPL}")
 
    excels = find_excels()
    if not excels:
        print("ℹ️ data/ 폴더에 엑셀(.xlsx)이 없어 빌드를 건너뜁니다.")
        return
 
    print(f"[1/4] 엑셀 {len(excels)}개 로드:")
    datasets = []
    for path in excels:
        header, rows = read_excel(path)
        label = tab_label(path.name)
        datasets.append({"label": label, "filename": path.name, "header": header, "rows": rows})
        print(f"      - {path.name} → 탭 '{label}' ({len(rows)}행)")
 
    print("[2/4] 내부 HTML 생성")
    inner_tpl = INNER_TPL.read_text(encoding="utf-8")
    data_json = json.dumps(datasets, ensure_ascii=False)
    inner = inner_tpl.replace("__DATA__", data_json)
    print(f"      내부 HTML 크기: {len(inner):,} bytes")
 
    print("[3/4] AES 암호화")
    d = encrypt(inner, password)
    print(f"      ciphertext: {len(d['ciphertext']):,} chars")
 
    print("[4/4] 래퍼에 주입 후 저장")
    wrapper = WRAPPER_TPL.read_text(encoding="utf-8")
    final = wrapper.replace("__D_PLACEHOLDER__", json.dumps(d, ensure_ascii=False))
    OUTPUT.write_text(final, encoding="utf-8")
    print(f"  -> {OUTPUT.name} ({OUTPUT.stat().st_size:,} bytes)")
    print("완료.")
 
 
if __name__ == "__main__":
    main()
